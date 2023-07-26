import logging
import os

import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

from config import Up_To_Date_NEWS_FILE
from excel_sheet import create_workbook, create_or_load_sheet, insert_rows


def make_links_clickable(sheet, source):
    """
    Make the links in the sheet clickable.

    Args:
        sheet (openpyxl.Worksheet): Sheet object.
        source (str): Source string representing the type of articles.

    Returns:
        None
    """
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
        for (i, cell) in enumerate(row, start=1):
            if (source.startswith('Daily-Updates') and i == 3) or (not source.startswith('Daily-Updates') and i == 2):
                cell.hyperlink = cell.value
                cell.style = 'Hyperlink'


def limit_articles(sheet, source):
    """
    Limit the number of articles in the sheet.

    Args:
        sheet (openpyxl.Worksheet): Sheet object.
        source (str): Source string representing the type of articles.

    Returns:
        None
    """
    if not source.startswith('Daily-Updates'):
        while sheet.max_row - 2 > 50:
            sheet.delete_rows(sheet.max_row)


def apply_border_font(sheet, source):
    """
    Apply the border and font to a specific sheet.

    Args:
        sheet (openpyxl.Worksheet): Sheet object.
        source (str): Source string representing the type of articles.

    Returns:
        None
    """
    if source.startswith('Daily-Updates'):
        add_borders(sheet)
        change_font_daily_updates(sheet)
    else:
        add_borders(sheet)
        change_font_author(sheet)


def append_to_excel(df, source):
    """
    Insert the DataFrame to the Excel file.

    Args:
        df (pandas.DataFrame): DataFrame object containing the articles.
        source (str): Source string representing the type of articles.

    Returns:
        openpyxl.Workbook: The modified Workbook object.
    """
    if os.path.exists(Up_To_Date_NEWS_FILE):
        book = openpyxl.load_workbook(Up_To_Date_NEWS_FILE)
    else:
        book = create_workbook()

    sheet = create_or_load_sheet(book, source)
    insert_rows(df, sheet)
    make_links_clickable(sheet, source)
    limit_articles(sheet, source)
    apply_border_font(sheet, source)

    for row in sheet.iter_rows(min_row=3, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    return book


def change_font_daily_updates(sheet):
    """
    Change the font of the titles and authors in the sheet for daily updates.

    Args:
        sheet (openpyxl.Worksheet): Sheet object.

    Returns:
        None
    """
    title_font = Font(name='Arial', size=14, bold=True)
    author_font = Font(name='Times New Roman', size=12, italic=True)

    try:
        for (i, row) in enumerate(sheet.iter_rows(min_row=3, min_col=1, max_col=3), start=3):
            for (j, cell) in enumerate(row, start=1):
                if j == 1:
                    cell.font = author_font
                elif j == 2:
                    cell.font = title_font

                if i % 2 == 0:
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                else:
                    cell.fill = PatternFill(fill_type=None)
    except Exception as e:
        logging.exception(f"Error while changing font in sheet '{sheet.title}': {e}")
        raise


def change_font_author(sheet):
    """
    Change the font of the titles in the sheet for author updates.

    Args:
        sheet (openpyxl.Worksheet): Sheet object.

    Returns:
        None
    """
    title_font = Font(name='Arial', size=14)
    date_font = Font(name='Consolas', size=12)

    try:
        for (i, row) in enumerate(sheet.iter_rows(min_row=3, min_col=1, max_col=3), start=3):
            for (j, cell) in enumerate(row, start=1):
                if j == 1:
                    cell.font = title_font
                elif j == 3:
                    cell.font = date_font

                if i % 2 == 0:
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                else:
                    cell.fill = PatternFill(fill_type=None)
    except Exception as e:
        logging.exception(f"Error while changing font in sheet '{sheet.title}': {e}")
        raise


def add_borders(sheet):
    """
    Add borders around the cells in the sheet.

    Args:
        sheet (openpyxl.Worksheet): Sheet object.

    Returns:
        None
    """
    A = 'thin'
    thin_border = Border(left=Side(style=A), right=Side(style=A), top=Side(style=A), bottom=Side(style=A))

    try:
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
    except Exception as e:
        logging.exception(f"Error while applying border to sheet '{sheet.title}': {e}")
        raise


def save_articles(articles, source):
    """
    Save articles to an Excel file.

    Args:
        articles (list): List of articles to be saved.
        source (str): Source string representing the type of articles.

    Returns:
        None
    """
    if source.startswith('Daily-Updates'):
        df = pd.DataFrame(articles, columns=['Source', 'Title', 'Link'])
    else:
        df = pd.DataFrame(articles, columns=['Title', 'Link', 'Date'])

    book = append_to_excel(df, source)
    book.save(Up_To_Date_NEWS_FILE)

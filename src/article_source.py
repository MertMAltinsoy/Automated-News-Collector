# article_source.py
import logging

from openpyxl.styles import Font, PatternFill, Alignment

from config import COLORS


def get_source_and_headers(source):
    """
    Get the color and headers for the source.

    Args:
        source (str): Source string representing the type of articles.

    Returns:
        Tuple: Tuple containing the source color and headers.
    """
    if source.startswith('Daily-Updates'):
        sourceColor = COLORS['Daily-Updates']
        headers = ['Author', 'Title', 'Link']
    else:
        sourceColor = COLORS.get(source, '000000')
        headers = ['Title', 'Link', 'Date']

    return sourceColor, headers


def add_source_header(sheet, source, sourceColor):
    """
    Add the source header to the sheet.

    Args:
        sheet (openpyxl.Worksheet): Sheet object.
        source (str): Source string representing the type of articles.
        sourceColor (str): String representing the color of the source.

    Returns:
        None
    """
    stripped_source = replace_dash_with_space(source)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    source_cell = sheet.cell(row=1, column=1, value=stripped_source)
    source_cell.font = Font(bold=True, color='FFFFFF', size=16)
    source_cell.fill = PatternFill(start_color=sourceColor, end_color=sourceColor, fill_type='solid')
    source_cell.alignment = Alignment(horizontal='center', vertical='center')


def add_headers(sheet, headers, headersColor):
    """
    Add the headers to the sheet.

    Args:
        sheet (openpyxl.Worksheet): Sheet object.
        headers (List[str]): List of strings representing the headers.
        headersColor (str): String representing the color of the headers.

    Returns:
        None
    """
    sheet.append(headers)
    for cell in sheet[2:2]:
        cell.font = Font(bold=True, color='FFFFFF', size=14)
        cell.fill = PatternFill(start_color=headersColor, end_color=headersColor, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.freeze_panes = 'A3'
    sheet.auto_filter.ref = f"A2:C{sheet.max_row}"


def replace_dash_with_space(source):
    """
    Replace dashes with spaces in the source string.

    Args:
        source (str): Source string representing the type of articles.

    Returns:
        str: String with dashes replaced by spaces.
    """
    try:
        if source.startswith('Daily-Updates'):
            return source.replace('-', ' ', 2)
        else:
            return source.replace('-', ' ')
    except Exception as e:
        logging.exception(f"Error replacing dashes with spaces in source '{source}': {e}")
        raise


def adjust_color(color, amount):
    """
    Adjust the brightness of a color.

    Args:
        color (str): String representing the color in hex format.
        amount (int): Integer representing the amount to adjust the brightness.

    Returns:
        str: String representing the adjusted color in hex format.
    """
    try:
        r, g, b = int(color[:2], 16), int(color[2:4], 16), int(color[4:], 16)
        r = max(0, min(255, r + amount))
        g = max(0, min(255, g + amount))
        b = max(0, min(255, b + amount))
        return '{:02x}{:02x}{:02x}'.format(r, g, b)
    except Exception as e:
        logging.exception(f"Error adjusting color '{color}': {e}")
        raise

import logging
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from article_source import add_headers, add_source_header, adjust_color, get_source_and_headers
from config import Up_To_Date_NEWS_FILE


def create_workbook():
    """
    Create a new Excel workbook.

    Returns:
        openpyxl.Workbook: Workbook object.
    """
    try:
        book = openpyxl.Workbook()
        del book['Sheet']
        return book
    except Exception as e:
        logging.exception(f"Error creating Excel workbook: {e}")
        raise


def create_sheet(book, source):
    """
    Create a new sheet with headers.

    Args:
        book: Workbook object.
        source: String representing the source of the data.

    Returns:
        Sheet object.
    """
    try:
        sheet = book.create_sheet(source)
        sourceColor, headers = get_source_and_headers(source)
        headersColor = adjust_color(sourceColor, -20)  # Make the headers color a bit darker
        add_source_header(sheet, source, sourceColor)
        add_headers(sheet, headers, headersColor)
        return sheet
    except Exception as e:
        logging.exception(f"Error creating sheet for source '{source}': {e}")
        raise


def create_or_load_sheet(book, source):
    """
    Create or load a sheet based on the source.

    Args:
        book (openpyxl.Workbook): Workbook object.
        source (str): Source string.

    Returns:
        openpyxl.Workbook: Sheet object.
    """
    if source in book.sheetnames:
        sheet = book[source]
    elif source.startswith('Daily-Updates'):
        sheet = create_sheet(book, source)
        book._sheets.sort(key=lambda sheet: sheet.title != source)
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 110
        sheet.column_dimensions['C'].width = 110
    else:
        sheet = create_sheet(book, source)
        sheet.column_dimensions['A'].width = 110
        sheet.column_dimensions['B'].width = 110
        sheet.column_dimensions['C'].width = 15

    return sheet


def insert_rows(df, sheet):
    """
    Insert rows from DataFrame to the sheet.

    Args:
        df (pandas.DataFrame): DataFrame object.
        sheet (openpyxl.Worksheet): Sheet object.

    Returns:
        None
    """
    for (index, row) in list(df.iterrows())[::-1]:
        sheet.insert_rows(3)
        for (i, value) in enumerate(row.tolist(), start=1):
            cell = sheet.cell(row=3, column=i, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')


def create_index_sheet():
    """
    Create or update an index sheet with links to all other sheets.

    Returns:
        None
    """
    try:
        if os.path.exists(Up_To_Date_NEWS_FILE):
            book = openpyxl.load_workbook(Up_To_Date_NEWS_FILE)
            if 'Index' in book.sheetnames:
                sheet = book['Index']  # Load the existing index sheet
                sheet.delete_rows(2, sheet.max_row)  # Delete all rows except the header
            else:
                sheet = book.create_sheet('Index', 1)  # Create a new sheet at the second position
                sheet.append(['Author'])  # Add headers
                sheet['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
                sheet['A1'].fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # Gray fill
                sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
                sheet.column_dimensions['A'].width = 60

            existing_authors = [cell.value for cell in sheet['A'] if cell.value != 'Author']
            for other_sheet in book.sheetnames[2:]:  # Skip the Index sheet itself
                if other_sheet not in existing_authors:
                    sheet.append([other_sheet])  # Add the name of the other sheet
                    link_cell = sheet.cell(row=sheet.max_row, column=1)
                    link_cell.value = f'=HYPERLINK("#\'{other_sheet}\'!A1", "{other_sheet}")'  # Add a formula that links to the other sheet
                    link_cell.style = 'Hyperlink'  # Make the cell look like a hyperlink
                    link_cell.font = Font(color='0000EE', underline='single', size=14)  # Blue, underlined text
                    link_cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Apply striping
                    if link_cell.row % 2 == 0:
                        link_cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Light gray fill

            # Apply border to the cells
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            book.save(Up_To_Date_NEWS_FILE)
        else:
            logging.info("Excel file does not exist yet. Index sheet will be created after the first run.")

    except Exception as e:
        logging.exception(f"Error creating or updating index sheet: {e}")
        raise
